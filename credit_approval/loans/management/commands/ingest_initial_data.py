from decimal import Decimal
from pathlib import Path

from django.conf import settings
from django.core.management.base import BaseCommand
from openpyxl import load_workbook

from ...models import Customer, Loan


def _to_decimal(value):
    if value is None:
        return Decimal("0")
    return Decimal(str(value))


def _to_date(value):
    if not value:
        return None
    return value.date() if hasattr(value, "date") else value


class Command(BaseCommand):
    help = "Load the provided customer and loan spreadsheets into the database"

    def handle(self, *args, **options):
        data_dir = Path(settings.BASE_DIR)
        customer_path = data_dir / "customer_data.xlsx"
        loan_path = data_dir / "loan_data.xlsx"
        if not customer_path.exists() or not loan_path.exists():
            self.stdout.write(self.style.ERROR("Data files not found in the project root."))
            return
        self._load_customers(customer_path)
        self._load_loans(loan_path)
        self.stdout.write(self.style.SUCCESS("Customer and loan data successfully ingested."))

    def _load_customers(self, file_path: Path) -> None:
        workbook = load_workbook(file_path)
        sheet = workbook.active
        rows = sheet.iter_rows(min_row=2, values_only=True)
        for row in rows:
            if not row or not row[0]:
                continue
            (
                customer_id,
                first_name,
                last_name,
                phone_number,
                monthly_salary,
                approved_limit,
                current_debt,
            ) = row
            Customer.objects.update_or_create(
                id=int(customer_id),
                defaults={
                    "first_name": str(first_name).strip(),
                    "last_name": str(last_name).strip(),
                    "phone_number": str(phone_number).strip(),
                    "age": 0,
                    "monthly_income": int(_to_decimal(monthly_salary)),
                    "approved_limit": _to_decimal(approved_limit),
                    "current_debt": _to_decimal(current_debt),
                },
            )

    def _load_loans(self, file_path: Path) -> None:
        workbook = load_workbook(file_path)
        sheet = workbook.active
        rows = sheet.iter_rows(min_row=2, values_only=True)
        for row in rows:
            if not row or not row[0]:
                continue
            (
                customer_id,
                loan_id,
                loan_amount,
                tenure,
                interest_rate,
                monthly_repayment,
                emis_paid_on_time,
                start_date,
                end_date,
            ) = row
            try:
                customer = Customer.objects.get(id=int(customer_id))
            except Customer.DoesNotExist:
                continue
            Loan.objects.update_or_create(
                id=int(loan_id),
                defaults={
                    "customer": customer,
                    "loan_amount": _to_decimal(loan_amount),
                    "tenure": int(tenure or 0),
                    "interest_rate": _to_decimal(interest_rate),
                    "monthly_installment": _to_decimal(monthly_repayment),
                    "emis_paid_on_time": int(emis_paid_on_time or 0),
                    "start_date": _to_date(start_date),
                    "end_date": _to_date(end_date),
                    "approved": True,
                },
            )
